import openpyxl
import ItemPub


"""
excel reader, reads inventory data from excel file and parses it into something useable
contains a strategy made for reading differnet types of excel files
supports 1 big sheet right now 
Author: Jack Frate
"""


class ExcelReader:
    def __init__(self, name: str):
        self.name = name
        wb = openpyxl.load_workbook(name)
        ls = wb.sheetnames
        self.sheet = wb[ls[0]]

    # method to parse a sheet and will return a dictionary of all of the items
    def parse_sheet(self) -> {}:
        ret = {}

        # now parse the sheet
        for i in range(1, self.sheet.max_row + 1):

            # checks to make sure that the column is not the header
            if self.sheet.cell(row=i, column=1).value == 'Name':
                continue

            # checks for duplicates, if there is one: continue to next loop and notify with a message
            if self.sheet.cell(row=i, column=1).value in ret:
                # outputs what row the duplicate is at
                print('Duplicate of item: ',
                      self.sheet.cell(row=i, column=1).value,
                      ' found, will not be recorded. \n Duplicate item found at row: ',
                      i)
                continue

            # checks for item having no category, and replaces it with none if there is not one
            if self.sheet.cell(row=i, column=2).value is None:
                cat = 'None'
            else:
                cat = self.sheet.cell(row=i, column=2).value

            # create the item for the dictionary listing
            item = ItemPub.ItemPub(self.sheet.cell(row=i, column=1).value,
                                   # variable cat checks for no category
                                   cat,
                                   self.sheet.cell(row=i, column=3).value,
                                   self.sheet.cell(row=i, column=4).value)

            # clean up the price
            item.clean_price()

            # now add the entry to the dictionary
            # entry is: ret['name of item'] = 'item object'
            ret[item.name] = item

        return ret

    # method that will print everything. Used for testing. TODO make into unit test
    def parse_sheet_list(self) -> {}:
        ret = {}
        ls = []

        # now parse the sheet
        for i in range(1, self.sheet.max_row + 1):

            # checks to make sure that the column is not the header
            if self.sheet.cell(row=i, column=1).value == 'Name':
                continue

            # checks for duplicates, if there is one: continue to next loop and notify with a message
            if self.sheet.cell(row=i, column=1).value in ret:
                # outputs what row the duplicate is at
                print('Duplicate of item: ',
                      self.sheet.cell(row=i, column=1).value,
                      ' found, will not be recorded. \n Duplicate item found at row: ',
                      i)
                continue

            # checks for item having no category, and replaces it with none if there is not one
            if self.sheet.cell(row=i, column=2).value is None:
                cat = 'None'
            else:
                cat = self.sheet.cell(row=i, column=2).value

            # create the item for the dictionary listing
            item = ItemPub.ItemPub(self.sheet.cell(row=i, column=1).value,
                                   # variable cat checks for no category
                                   cat,
                                   self.sheet.cell(row=i, column=3).value,
                                   self.sheet.cell(row=i, column=4).value)

            # now add the entry to the dictionary
            # entry is: ret['name of item'] = 'item object'
            ret[item.name] = item
            ls.append(item)

        return ls
